#!/usr/bin/env python3
"""
Canvas Data Export Module

Handles exporting Canvas data to various formats including CSV, JSON, and Excel.
Provides flexible export options with customizable field selection and formatting.
"""

import csv
import json
import os
import logging
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("pandas not available - Excel export will be limited")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel export will use basic format")

logger = logging.getLogger(__name__)

class CanvasExporter:
    """
    Export Canvas data to various formats
    """
    
    def __init__(self, output_directory: str = "./exports"):
        """
        Initialize exporter
        
        Args:
            output_directory: Directory to save exported files
        """
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(exist_ok=True)
        
        # Default field mappings for different data types
        self.field_mappings = {
            'courses': {
                'id': 'Course ID',
                'name': 'Course Name',
                'course_code': 'Course Code',
                'workflow_state': 'Status',
                'start_at': 'Start Date',
                'end_at': 'End Date',
                'enrollment_term_id': 'Term ID',
                'term_name': 'Term Name',
                'user_count': 'Enrolled Students',
                'assignment_count': 'Total Assignments'
            },
            'assignments': {
                'id': 'Assignment ID',
                'name': 'Assignment Name',
                'course_id': 'Course ID',
                'due_at': 'Due Date',
                'points_possible': 'Points Possible',
                'submission_types': 'Submission Types',
                'grading_type': 'Grading Type',
                'is_overdue': 'Is Overdue',
                'days_until_due': 'Days Until Due'
            },
            'grades': {
                'user_id': 'Student ID',
                'user_name': 'Student Name',
                'user_email': 'Email',
                'current_score': 'Current Score',
                'final_score': 'Final Score',
                'current_grade': 'Current Grade',
                'final_grade': 'Final Grade',
                'enrollment_state': 'Enrollment Status',
                'last_activity_at': 'Last Activity'
            },
            'users': {
                'id': 'User ID',
                'name': 'Full Name',
                'email': 'Email',
                'login_id': 'Login ID',
                'enrollments': 'Enrollments'
            }
        }
    
    def export_to_csv(self, data: List[Dict], filename: str, data_type: str = None, 
                     custom_fields: List[str] = None) -> str:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            data_type: Type of data for field mapping
            custom_fields: Custom field selection
        
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"{filename}_{timestamp}.csv"
        csv_path = self.output_directory / csv_filename
        
        try:
            # Determine fields to export
            if custom_fields:
                fields = custom_fields
            elif data_type and data_type in self.field_mappings:
                fields = list(self.field_mappings[data_type].keys())
            else:
                # Use all available fields from first record
                fields = list(data[0].keys())
            
            # Get field headers
            headers = self._get_headers(fields, data_type)
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fields, extrasaction='ignore')
                
                # Write headers
                writer.writerow(dict(zip(fields, headers)))
                
                # Write data
                for row in data:
                    # Clean and format row data
                    cleaned_row = self._clean_row_data(row, fields)
                    writer.writerow(cleaned_row)
            
            logger.info(f"Exported {len(data)} records to {csv_path}")
            return str(csv_path)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_to_json(self, data: Union[List[Dict], Dict], filename: str, 
                      pretty_print: bool = True) -> str:
        """
        Export data to JSON format
        
        Args:
            data: Data to export (list of dicts or single dict)
            filename: Output filename (without extension)
            pretty_print: Whether to format JSON for readability
        
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_filename = f"{filename}_{timestamp}.json"
        json_path = self.output_directory / json_filename
        
        try:
            # Prepare export data with metadata
            export_data = {
                'exported_at': datetime.now().isoformat(),
                'record_count': len(data) if isinstance(data, list) else 1,
                'data': data
            }
            
            with open(json_path, 'w', encoding='utf-8') as jsonfile:
                if pretty_print:
                    json.dump(export_data, jsonfile, indent=2, default=str, ensure_ascii=False)
                else:
                    json.dump(export_data, jsonfile, default=str, ensure_ascii=False)
            
            record_count = len(data) if isinstance(data, list) else 1
            logger.info(f"Exported {record_count} records to {json_path}")
            return str(json_path)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise
    
    def export_to_excel(self, data: Union[List[Dict], Dict[str, List[Dict]]], 
                       filename: str, data_type: str = None) -> str:
        """
        Export data to Excel format
        
        Args:
            data: Data to export (list of dicts or dict with sheet names as keys)
            filename: Output filename (without extension)
            data_type: Type of data for field mapping
        
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{filename}_{timestamp}.xlsx"
        excel_path = self.output_directory / excel_filename
        
        try:
            if PANDAS_AVAILABLE and OPENPYXL_AVAILABLE:
                return self._export_to_excel_advanced(data, excel_path, data_type)
            else:
                return self._export_to_excel_basic(data, excel_path, data_type)
                
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _export_to_excel_advanced(self, data: Union[List[Dict], Dict[str, List[Dict]]], 
                                 excel_path: Path, data_type: str = None) -> str:
        """Export to Excel using pandas and openpyxl for advanced formatting"""
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            if isinstance(data, dict):
                # Multiple sheets
                for sheet_name, sheet_data in data.items():
                    if sheet_data:
                        df = pd.DataFrame(sheet_data)
                        df = self._format_dataframe(df, data_type)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._format_excel_sheet(writer.sheets[sheet_name], df)
            else:
                # Single sheet
                df = pd.DataFrame(data)
                df = self._format_dataframe(df, data_type)
                sheet_name = data_type.title() if data_type else 'Data'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_excel_sheet(writer.sheets[sheet_name], df)
        
        record_count = len(data) if isinstance(data, list) else sum(len(v) for v in data.values())
        logger.info(f"Exported {record_count} records to {excel_path}")
        return str(excel_path)
    
    def _export_to_excel_basic(self, data: Union[List[Dict], Dict[str, List[Dict]]], 
                              excel_path: Path, data_type: str = None) -> str:
        """Basic Excel export without advanced formatting"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        if isinstance(data, dict):
            # Multiple sheets
            for sheet_name, sheet_data in data.items():
                if sheet_data:
                    ws = wb.create_sheet(title=sheet_name)
                    self._write_sheet_data(ws, sheet_data, data_type)
        else:
            # Single sheet
            sheet_name = data_type.title() if data_type else 'Data'
            ws = wb.create_sheet(title=sheet_name)
            self._write_sheet_data(ws, data, data_type)
        
        wb.save(excel_path)
        
        record_count = len(data) if isinstance(data, list) else sum(len(v) for v in data.values())
        logger.info(f"Exported {record_count} records to {excel_path}")
        return str(excel_path)
    
    def _write_sheet_data(self, worksheet, data: List[Dict], data_type: str = None):
        """Write data to Excel worksheet"""
        if not data:
            return
        
        # Determine fields and headers
        if data_type and data_type in self.field_mappings:
            fields = list(self.field_mappings[data_type].keys())
        else:
            fields = list(data[0].keys())
        
        headers = self._get_headers(fields, data_type)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            if OPENPYXL_AVAILABLE:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            cleaned_row = self._clean_row_data(row_data, fields)
            for col_idx, field in enumerate(fields, 1):
                value = cleaned_row.get(field, '')
                worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    def _format_dataframe(self, df: 'pd.DataFrame', data_type: str = None) -> 'pd.DataFrame':
        """Format DataFrame for better Excel output"""
        # Rename columns using field mappings
        if data_type and data_type in self.field_mappings:
            mapping = self.field_mappings[data_type]
            df = df.rename(columns={k: v for k, v in mapping.items() if k in df.columns})
        
        # Format datetime columns
        for col in df.columns:
            if df[col].dtype == 'object':
                # Try to convert datetime strings
                try:
                    df[col] = pd.to_datetime(df[col], errors='ignore')
                except:
                    pass
        
        return df
    
    def _format_excel_sheet(self, worksheet, df: 'pd.DataFrame'):
        """Apply formatting to Excel worksheet"""
        if not OPENPYXL_AVAILABLE:
            return
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    def _get_headers(self, fields: List[str], data_type: str = None) -> List[str]:
        """Get display headers for fields"""
        if data_type and data_type in self.field_mappings:
            mapping = self.field_mappings[data_type]
            return [mapping.get(field, field.replace('_', ' ').title()) for field in fields]
        else:
            return [field.replace('_', ' ').title() for field in fields]
    
    def _clean_row_data(self, row: Dict, fields: List[str]) -> Dict:
        """Clean and format row data for export"""
        cleaned = {}
        
        for field in fields:
            value = row.get(field)
            
            # Handle different data types
            if value is None:
                cleaned[field] = ''
            elif isinstance(value, datetime):
                cleaned[field] = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, (list, dict)):
                cleaned[field] = json.dumps(value) if value else ''
            elif isinstance(value, bool):
                cleaned[field] = 'Yes' if value else 'No'
            else:
                cleaned[field] = str(value)
        
        return cleaned
    
    def export_course_analytics(self, analytics_data: Dict, course_name: str) -> Dict[str, str]:
        """
        Export comprehensive course analytics to multiple formats
        
        Args:
            analytics_data: Course analytics data
            course_name: Name of the course for filename
        
        Returns:
            Dictionary with paths to exported files
        """
        # Clean course name for filename
        safe_course_name = "".join(c for c in course_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_course_name = safe_course_name.replace(' ', '_')
        
        exported_files = {}
        
        try:
            # Export to JSON (full analytics)
            json_path = self.export_to_json(analytics_data, f"course_analytics_{safe_course_name}")
            exported_files['json'] = json_path
            
            # Export to Excel (multiple sheets)
            excel_data = {
                'Course_Info': [analytics_data.get('course_info', {})],
                'Assignment_Analytics': [analytics_data.get('assignment_analytics', {})],
                'Grade_Analytics': [analytics_data.get('grade_analytics', {})]
            }
            
            excel_path = self.export_to_excel(excel_data, f"course_analytics_{safe_course_name}")
            exported_files['excel'] = excel_path
            
            logger.info(f"Exported course analytics for '{course_name}' to multiple formats")
            return exported_files
            
        except Exception as e:
            logger.error(f"Error exporting course analytics: {e}")
            raise
    
    def get_export_summary(self) -> Dict:
        """Get summary of exported files"""
        export_files = list(self.output_directory.glob('*'))
        
        summary = {
            'export_directory': str(self.output_directory),
            'total_files': len(export_files),
            'files_by_type': {},
            'recent_exports': []
        }
        
        # Count files by type
        for file_path in export_files:
            if file_path.is_file():
                ext = file_path.suffix.lower()
                summary['files_by_type'][ext] = summary['files_by_type'].get(ext, 0) + 1
                
                # Add to recent exports (last 10)
                if len(summary['recent_exports']) < 10:
                    summary['recent_exports'].append({
                        'filename': file_path.name,
                        'size_bytes': file_path.stat().st_size,
                        'modified': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
                    })
        
        # Sort recent exports by modification time
        summary['recent_exports'].sort(key=lambda x: x['modified'], reverse=True)
        
        return summary